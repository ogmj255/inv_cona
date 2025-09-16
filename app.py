from flask import Flask, render_template, request, redirect, url_for, session, flash, make_response
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from pymongo import MongoClient
from datetime import datetime, timedelta
from bson.objectid import ObjectId
from werkzeug.utils import secure_filename
import re
import os
import bcrypt
import base64
from io import BytesIO
import logging
from functools import wraps

# Importar librerías para reportes y QR
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import qrcode
    from PIL import Image
    QR_AVAILABLE = True
except ImportError:
    QR_AVAILABLE = False

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'a1eb8b7d4c7a96ea202923296486a51c')

@app.context_processor
def inject_datetime():
    return {'datetime': datetime}

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# MongoDB connection
mongo_uri = os.environ.get('MONGODB_URI', 'mongodb+srv://ogmoscosoj2:mQaZ63iaQO7feFXo@conagoparedbinventario.ecb0dj0.mongodb.net/?retryWrites=true&w=majority&appName=conagoparedbinventario')
local_uri = 'mongodb://localhost:27017/cona_inv'

try:
    client = MongoClient(
        mongo_uri,
        serverSelectionTimeoutMS=10000,
        connectTimeoutMS=10000,
        socketTimeoutMS=10000,
        maxPoolSize=10,
        retryWrites=True
    )
    client.admin.command('ping')
    print("Conectado exitosamente a MongoDB Atlas")
except Exception as e:
    print(f"Error conectando a MongoDB Atlas: {e}")
    try:
        client = MongoClient(local_uri, serverSelectionTimeoutMS=5000)
        client.admin.command('ping')
        print("Conectado exitosamente a MongoDB local")
    except Exception as local_e:
        print(f"Error conectando a MongoDB local: {local_e}")
        exit(1)

cona_inv = client['cona_inv']
users = cona_inv['users']
parroquias = cona_inv['parroquias']
inventarios = cona_inv['inventarios']
bienes_asignados = cona_inv['bienes_asignados']
audit_logs = cona_inv['audit_logs']
notificaciones = cona_inv['notificaciones']

def log_action(action, details, user_id=None):
    """Registrar acciones en el sistema de auditoría"""
    try:
        log_entry = {
            'action': action,
            'details': details,
            'user_id': user_id or (current_user.username if current_user.is_authenticated else 'anonymous'),
            'timestamp': datetime.now(),
            'ip_address': request.remote_addr if request else 'unknown'
        }
        audit_logs.insert_one(log_entry)
    except Exception as e:
        print(f"Error logging action: {e}")

def create_notification(user_id, message, tipo='info'):
    """Crear notificación para usuario"""
    try:
        notification = {
            'user_id': user_id,
            'message': message,
            'tipo': tipo,
            'leida': False,
            'timestamp': datetime.now()
        }
        notificaciones.insert_one(notification)
    except Exception as e:
        print(f"Error creating notification: {e}")

class User(UserMixin):
    def __init__(self, username, role, parroquia_id=None):
        self.id = username 
        self.username = username
        self.role = role
        self.parroquia_id = parroquia_id

@login_manager.user_loader
def load_user(username):
    try:
        user_data = users.find_one({'username': username})
        if user_data:
            return User(username=user_data['username'], role=user_data['role'], parroquia_id=user_data.get('parroquia_id'))
        return None
    except Exception as e:
        print(f"Error loading user: {e}")
        return None

@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role == 'super_admin':
            return redirect(url_for('super_admin'))
        elif current_user.role == 'admin_parroquia':
            return redirect(url_for('admin_parroquia'))
        elif current_user.role == 'tecnico':
            return redirect(url_for('admin_parroquia'))
    return render_template('login.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = users.find_one({'username': username})
        if user and bcrypt.checkpw(password.encode('utf-8'), user['password'].encode('utf-8') if isinstance(user['password'], str) else user['password']):
            user_obj = User(username=user['username'], role=user['role'], parroquia_id=user.get('parroquia_id'))
            login_user(user_obj)
            session['full_name'] = f"{user.get('nombre', '')} {user.get('apellido', '')}".strip() or username
            log_action('LOGIN', f'Usuario {username} inició sesión', username)
            return redirect(url_for('index'))
        flash('Usuario o contraseña incorrectos', 'error')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    session.clear()
    return redirect(url_for('login'))

@app.route('/super_admin')
@app.route('/super_admin/<section>', methods=['GET', 'POST'])
@login_required
def super_admin(section=None):
    if current_user.role != 'super_admin':
        return redirect(url_for('index'))
    
    # Calcular bienes por tipo
    pipeline = [
        {'$group': {'_id': '$tipo', 'count': {'$sum': 1}}},
        {'$sort': {'count': -1}}
    ]
    tipos_result = list(inventarios.aggregate(pipeline))
    bienes_por_tipo = {}
    for tipo in tipos_result:
        if tipo['_id']:
            bienes_por_tipo[tipo['_id']] = tipo['count']
    
    stats = {
        'total_parroquias': parroquias.count_documents({}),
        'total_usuarios': users.count_documents({}),
        'total_bienes': inventarios.count_documents({}),
        'bienes_asignados': bienes_asignados.count_documents({'estado': 'activo'}),
        'total_tecnicos': users.count_documents({'role': 'tecnico'}),
        'asignaciones_activas': bienes_asignados.count_documents({'estado': 'activo'}),
        'asignaciones_devueltas': bienes_asignados.count_documents({'estado': 'devuelto'}),
        'parroquias_con_bienes': parroquias.count_documents({}),
        'responsables_activos': users.count_documents({'role': {'$in': ['admin_parroquia', 'tecnico']}}),
        'bienes_por_estado': {
            'disponibles': inventarios.count_documents({'estado': 'disponible'}),
            'asignados': inventarios.count_documents({'estado': 'asignado'}),
            'mantenimiento': inventarios.count_documents({'estado': 'en_mantenimiento'}),
            'dañados': inventarios.count_documents({'estado': 'dañado'})
        },
        'usuarios_por_rol': {
            'super_admin': users.count_documents({'role': 'super_admin'}),
            'admin_parroquia': users.count_documents({'role': 'admin_parroquia'}),
            'tecnico': users.count_documents({'role': 'tecnico'})
        },
        'bienes_por_tipo': bienes_por_tipo
    }
    
    if section == 'parroquias':
        parroquias_list = list(parroquias.find())
        return render_template('gestionar_parroquias.html', stats=stats, parroquias=parroquias_list)
    elif section == 'usuarios':
        usuarios_list = list(users.find())
        parroquias_list = list(parroquias.find())
        return render_template('gestionar_usuarios.html', stats=stats, usuarios=usuarios_list, parroquias=parroquias_list)
    elif section == 'estadisticas':
        return render_template('estadisticas.html', stats=stats, estadisticas=stats)
    elif section == 'auditoria':
        logs = list(audit_logs.find().sort('timestamp', -1).limit(100))
        return render_template('auditoria.html', stats=stats, logs=logs)
    elif section == 'mantenimiento':
        return render_template('mantenimiento.html', stats=stats)
    
    return render_template('super_admin.html', stats=stats)

@app.route('/admin_parroquia')
@app.route('/admin_parroquia/<section>', methods=['GET', 'POST'])
@login_required
def admin_parroquia(section=None):
    if current_user.role not in ['admin_parroquia', 'tecnico']:
        return redirect(url_for('index'))
    
    # Redirigir técnicos a su panel específico
    if current_user.role == 'tecnico':
        return redirect(url_for('panel_tecnico'))
    
    # Buscar parroquia del usuario o crear una temporal
    parroquia_info = None
    if current_user.parroquia_id:
        parroquia_info = parroquias.find_one({'_id': ObjectId(current_user.parroquia_id)})
    
    if not parroquia_info:
        parroquia_info = {'nombre': 'Parroquia Demo', 'canton': 'Canton Demo'}
    
    # Calcular estadísticas correctas
    if current_user.parroquia_id:
        parroquia_filter = {'parroquia_id': ObjectId(current_user.parroquia_id)}
        stats = {
            'total_bienes': inventarios.count_documents(parroquia_filter),
            'bienes_disponibles': inventarios.count_documents({**parroquia_filter, 'estado': 'disponible'}),
            'bienes_asignados': inventarios.count_documents({**parroquia_filter, 'estado': 'asignado'}),
            'bienes_mantenimiento': inventarios.count_documents({**parroquia_filter, 'estado': 'en_mantenimiento'}),
            'total_tecnicos': users.count_documents({'role': 'tecnico', 'parroquia_id': ObjectId(current_user.parroquia_id)})
        }
    else:
        stats = {
            'total_bienes': inventarios.count_documents({}),
            'bienes_disponibles': inventarios.count_documents({'estado': 'disponible'}),
            'bienes_asignados': inventarios.count_documents({'estado': 'asignado'}),
            'bienes_mantenimiento': inventarios.count_documents({'estado': 'en_mantenimiento'}),
            'total_tecnicos': users.count_documents({'role': 'tecnico'})
        }
    
    if section == 'inventario':
        # Filtrar por parroquia si no es super admin
        if current_user.role == 'super_admin':
            bienes_list = list(inventarios.find())
        else:
            filter_query = {'parroquia_id': ObjectId(current_user.parroquia_id)} if current_user.parroquia_id else {}
            bienes_list = list(inventarios.find(filter_query))
        return render_template('gestionar_inventario.html', parroquia=parroquia_info, stats=stats, bienes=bienes_list)
    elif section == 'asignaciones':
        return redirect(url_for('gestionar_asignaciones'))
    elif section == 'tecnicos':
        return redirect(url_for('gestionar_tecnicos'))
    elif section == 'estadisticas':
        # Calcular bienes por tipo para la parroquia
        if current_user.parroquia_id:
            pipeline = [
                {'$match': {'parroquia_id': ObjectId(current_user.parroquia_id)}},
                {'$group': {'_id': '$tipo', 'count': {'$sum': 1}}},
                {'$sort': {'count': -1}}
            ]
        else:
            pipeline = [
                {'$group': {'_id': '$tipo', 'count': {'$sum': 1}}},
                {'$sort': {'count': -1}}
            ]
        
        tipos_result = list(inventarios.aggregate(pipeline))
        bienes_por_tipo = []
        for tipo in tipos_result:
            if tipo['_id']:
                bienes_por_tipo.append({
                    '_id': tipo['_id'],
                    'cantidad': tipo['count']
                })
        
        # Obtener técnicos con sus bienes asignados
        tecnicos_con_bienes = []
        if current_user.parroquia_id:
            asignaciones_activas = list(bienes_asignados.find({
                'estado': 'activo',
                'parroquia_id': ObjectId(current_user.parroquia_id)
            }))
        else:
            asignaciones_activas = list(bienes_asignados.find({'estado': 'activo'}))
        
        # Agrupar por técnico
        tecnicos_dict = {}
        for asignacion in asignaciones_activas:
            tecnico_info = users.find_one({'_id': asignacion.get('tecnico_id')})
            bien_info = inventarios.find_one({'_id': asignacion.get('bien_id')})
            
            if tecnico_info and bien_info:
                tecnico_key = str(asignacion.get('tecnico_id'))
                if tecnico_key not in tecnicos_dict:
                    tecnicos_dict[tecnico_key] = {
                        'nombre': f"{tecnico_info.get('nombre', '')} {tecnico_info.get('apellido', '')}".strip(),
                        'bienes': []
                    }
                tecnicos_dict[tecnico_key]['bienes'].append({
                    'codigo': bien_info.get('codigo', ''),
                    'nombre': bien_info.get('nombre', ''),
                    'tipo': bien_info.get('tipo', '')
                })
        
        tecnicos_con_bienes = list(tecnicos_dict.values())
        
        stats['bienes_por_tipo'] = bienes_por_tipo
        stats['tecnicos_con_bienes'] = tecnicos_con_bienes
        return render_template('estadisticas_parroquia.html', parroquia=parroquia_info, stats=stats, estadisticas=stats)
    
    return render_template('admin_parroquia.html', parroquia=parroquia_info, stats=stats)

@app.route('/gestionar_parroquias', methods=['GET', 'POST'])
@login_required
def gestionar_parroquias():
    if current_user.role != 'super_admin':
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        if 'create_parroquia' in request.form:
            nombre = request.form.get('nombre', '').strip()
            canton = request.form.get('canton', '').strip()
            codigo = request.form.get('codigo', '').strip()
            
            if not all([nombre, canton, codigo]):
                flash('Todos los campos son requeridos', 'error')
                return redirect(url_for('gestionar_parroquias'))
            
            # Verificar código único
            if parroquias.find_one({'codigo': codigo}):
                flash('El código ya existe', 'error')
                return redirect(url_for('gestionar_parroquias'))
            
            parroquia = {
                'nombre': nombre,
                'canton': canton,
                'codigo': codigo,
                'created_at': datetime.now()
            }
            
            try:
                parroquias.insert_one(parroquia)
                log_action('CREATE_PARROQUIA', f'Parroquia {nombre} creada')
                flash('Parroquia creada exitosamente', 'success')
            except Exception as e:
                flash(f'Error al crear parroquia: {str(e)}', 'error')
        
        elif 'edit_parroquia' in request.form:
            parroquia_id = request.form.get('parroquia_id')
            if not parroquia_id:
                flash('ID de parroquia inválido', 'error')
                return redirect(url_for('gestionar_parroquias'))
            
            try:
                update_data = {
                    'nombre': request.form.get('edit_nombre', '').strip(),
                    'canton': request.form.get('edit_canton', '').strip(),
                    'codigo': request.form.get('edit_codigo', '').strip()
                }
                
                if not all(update_data.values()):
                    flash('Todos los campos son requeridos', 'error')
                    return redirect(url_for('gestionar_parroquias'))
                
                parroquias.update_one({'_id': ObjectId(parroquia_id)}, {'$set': update_data})
                log_action('UPDATE_PARROQUIA', f'Parroquia {update_data["nombre"]} actualizada')
                flash('Parroquia actualizada exitosamente', 'success')
            except Exception as e:
                flash(f'Error al actualizar parroquia: {str(e)}', 'error')
        
        elif 'delete_parroquia' in request.form:
            parroquia_id = request.form.get('parroquia_id')
            if not parroquia_id:
                flash('ID de parroquia inválido', 'error')
                return redirect(url_for('gestionar_parroquias'))
            
            try:
                # Verificar si hay usuarios asignados
                if users.find_one({'parroquia_id': ObjectId(parroquia_id)}):
                    flash('No se puede eliminar: hay usuarios asignados a esta parroquia', 'error')
                    return redirect(url_for('gestionar_parroquias'))
                
                parroquia = parroquias.find_one({'_id': ObjectId(parroquia_id)})
                parroquias.delete_one({'_id': ObjectId(parroquia_id)})
                log_action('DELETE_PARROQUIA', f'Parroquia {parroquia.get("nombre", "")} eliminada')
                flash('Parroquia eliminada exitosamente', 'success')
            except Exception as e:
                flash(f'Error al eliminar parroquia: {str(e)}', 'error')
    
    parroquias_list = list(parroquias.find())
    stats = {
        'total_parroquias': parroquias.count_documents({}),
        'total_usuarios': users.count_documents({}),
        'total_bienes': inventarios.count_documents({}),
        'bienes_asignados': bienes_asignados.count_documents({'estado': 'activo'}),
        'total_tecnicos': users.count_documents({'role': 'tecnico'})
    }
    return render_template('gestionar_parroquias.html', parroquias=parroquias_list, stats=stats)

@app.route('/gestionar_usuarios', methods=['GET', 'POST'])
@login_required
def gestionar_usuarios():
    if current_user.role != 'super_admin':
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        if 'create_user' in request.form:
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '')
            nombre = request.form.get('nombre', '').strip()
            apellido = request.form.get('apellido', '').strip()
            email = request.form.get('email', '').strip()
            role = request.form.get('role', '')
            parroquia_id = request.form.get('parroquia_id')
            
            # Validaciones
            if not all([username, password, nombre, apellido, email, role]):
                flash('Todos los campos son requeridos', 'error')
                return redirect(url_for('gestionar_usuarios'))
            
            if len(password) < 6:
                flash('La contraseña debe tener al menos 6 caracteres', 'error')
                return redirect(url_for('gestionar_usuarios'))
            
            # Verificar username único
            if users.find_one({'username': username}):
                flash('El nombre de usuario ya existe', 'error')
                return redirect(url_for('gestionar_usuarios'))
            
            # Verificar email único
            if users.find_one({'email': email}):
                flash('El email ya está registrado', 'error')
                return redirect(url_for('gestionar_usuarios'))
            
            hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
            
            usuario = {
                'username': username,
                'password': hashed_password,
                'nombre': nombre,
                'apellido': apellido,
                'email': email,
                'role': role,
                'parroquia_id': ObjectId(parroquia_id) if parroquia_id else None,
                'created_at': datetime.now()
            }
            
            try:
                users.insert_one(usuario)
                log_action('CREATE_USER', f'Usuario {username} creado con rol {role}')
                flash('Usuario creado exitosamente', 'success')
            except Exception as e:
                flash(f'Error al crear usuario: {str(e)}', 'error')
        
        elif 'edit_user' in request.form:
            user_id = request.form['user_id']
            update_data = {
                'nombre': request.form['edit_nombre'],
                'apellido': request.form['edit_apellido'],
                'username': request.form['edit_username'],
                'email': request.form['edit_email'],
                'role': request.form['edit_role'],
                'parroquia_id': ObjectId(request.form['edit_parroquia_id']) if request.form.get('edit_parroquia_id') else None
            }
            if request.form.get('edit_password'):
                update_data['password'] = bcrypt.hashpw(request.form['edit_password'].encode('utf-8'), bcrypt.gensalt())
            
            try:
                users.update_one({'_id': ObjectId(user_id)}, {'$set': update_data})
                flash('Usuario actualizado exitosamente', 'success')
            except Exception as e:
                flash(f'Error al actualizar usuario: {str(e)}', 'error')
        
        elif 'delete_user' in request.form:
            user_id = request.form['user_id']
            try:
                users.delete_one({'_id': ObjectId(user_id)})
                flash('Usuario eliminado exitosamente', 'success')
            except Exception as e:
                flash(f'Error al eliminar usuario: {str(e)}', 'error')
    
    usuarios_list = list(users.find())
    parroquias_list = list(parroquias.find())
    stats = {
        'total_parroquias': parroquias.count_documents({}),
        'total_usuarios': users.count_documents({}),
        'total_bienes': inventarios.count_documents({}),
        'bienes_asignados': bienes_asignados.count_documents({'estado': 'activo'}),
        'total_tecnicos': users.count_documents({'role': 'tecnico'})
    }
    return render_template('gestionar_usuarios.html', usuarios=usuarios_list, parroquias=parroquias_list, stats=stats)

@app.route('/gestionar_inventario', methods=['GET', 'POST'])
@login_required
def gestionar_inventario():
    if request.method == 'POST':
        if 'edit_bien' in request.form:
            bien_id = request.form.get('bien_id')
            if bien_id:
                try:
                    update_data = {
                        'codigo': request.form.get('edit_codigo', '').strip(),
                        'nombre': request.form.get('edit_nombre', '').strip(),
                        'tipo': request.form.get('edit_tipo', '').strip(),
                        'marca': request.form.get('edit_marca', '').strip(),
                        'modelo': request.form.get('edit_modelo', '').strip(),
                        'descripcion': request.form.get('edit_descripcion', '')
                    }
                    inventarios.update_one({'_id': ObjectId(bien_id)}, {'$set': update_data})
                    log_action('UPDATE_BIEN', f'Bien {update_data["codigo"]} actualizado')
                    flash('Bien actualizado exitosamente', 'success')
                except Exception as e:
                    flash(f'Error al actualizar bien: {str(e)}', 'error')
        elif 'delete_bien' in request.form:
            bien_id = request.form.get('bien_id')
            if bien_id:
                try:
                    inventarios.delete_one({'_id': ObjectId(bien_id)})
                    log_action('DELETE_BIEN', f'Bien eliminado')
                    flash('Bien eliminado exitosamente', 'success')
                except Exception as e:
                    flash(f'Error al eliminar bien: {str(e)}', 'error')
        elif 'add_bien' in request.form:
            codigo = request.form.get('codigo', '').strip()
            nombre = request.form.get('nombre', '').strip()
            tipo = request.form.get('tipo', '').strip()
            marca = request.form.get('marca', '').strip()
            modelo = request.form.get('modelo', '').strip()
            estado = 'disponible'
            descripcion = request.form.get('descripcion', '')
            
            if not all([codigo, nombre, tipo, marca, modelo]):
                flash('Todos los campos obligatorios son requeridos', 'error')
                return redirect(url_for('gestionar_inventario'))
            
            if inventarios.find_one({'codigo': codigo}):
                flash('El código ya existe', 'error')
                return redirect(url_for('gestionar_inventario'))
            
            bien = {
                'codigo': codigo,
                'nombre': nombre,
                'tipo': tipo,
                'marca': marca,
                'modelo': modelo,
                'estado': estado,
                'descripcion': descripcion,
                'parroquia_id': ObjectId(current_user.parroquia_id) if current_user.parroquia_id else None,
                'created_at': datetime.now()
            }
            
            try:
                inventarios.insert_one(bien)
                log_action('CREATE_BIEN', f'Bien {codigo} - {nombre} creado')
                flash('Bien creado exitosamente', 'success')
            except Exception as e:
                flash(f'Error al crear bien: {str(e)}', 'error')
    
    # Filtrar por parroquia si no es super admin
    if current_user.role == 'super_admin':
        bienes_list = list(inventarios.find())
        parroquia_info = {'nombre': 'Sistema CONA', 'canton': 'Ecuador'}
    else:
        filter_query = {'parroquia_id': ObjectId(current_user.parroquia_id)} if current_user.parroquia_id else {}
        bienes_list = list(inventarios.find(filter_query))
        parroquia_info = parroquias.find_one({'_id': ObjectId(current_user.parroquia_id)}) if current_user.parroquia_id else {'nombre': 'Sistema CONA', 'canton': 'Ecuador'}
    
    stats = {
        'total_bienes': len(bienes_list),
        'bienes_disponibles': len([b for b in bienes_list if b.get('estado') == 'disponible']),
        'bienes_asignados': len([b for b in bienes_list if b.get('estado') == 'asignado']),
        'bienes_mantenimiento': len([b for b in bienes_list if b.get('estado') == 'en_mantenimiento']),
        'total_tecnicos': users.count_documents({'role': 'tecnico'})
    }
    return render_template('gestionar_inventario.html', bienes=bienes_list, parroquia=parroquia_info, stats=stats)

@app.route('/gestionar_tecnicos', methods=['GET', 'POST'])
@login_required
def gestionar_tecnicos():
    if request.method == 'POST':
        if 'create_tecnico' in request.form:
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '')
            nombre = request.form.get('nombre', '').strip()
            apellido = request.form.get('apellido', '').strip()
            email = request.form.get('email', '').strip()
            
            # Validaciones
            if not all([username, password, nombre, apellido, email]):
                flash('Todos los campos son requeridos', 'error')
                return redirect(url_for('gestionar_tecnicos'))
            
            if len(password) < 6:
                flash('La contraseña debe tener al menos 6 caracteres', 'error')
                return redirect(url_for('gestionar_tecnicos'))
            
            # Verificar username único
            if users.find_one({'username': username}):
                flash('El nombre de usuario ya existe', 'error')
                return redirect(url_for('gestionar_tecnicos'))
            
            # Verificar email único
            if users.find_one({'email': email}):
                flash('El email ya está registrado', 'error')
                return redirect(url_for('gestionar_tecnicos'))
            
            hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
            
            tecnico = {
                'username': username,
                'password': hashed_password,
                'nombre': nombre,
                'apellido': apellido,
                'email': email,
                'role': 'tecnico',
                'parroquia_id': ObjectId(current_user.parroquia_id) if current_user.parroquia_id else None,
                'created_at': datetime.now()
            }
            
            try:
                result = users.insert_one(tecnico)
                if result.inserted_id:
                    log_action('CREATE_TECNICO', f'Técnico {username} creado para parroquia {current_user.parroquia_id}')
                    flash(f'Técnico {nombre} {apellido} creado exitosamente', 'success')
                else:
                    flash('Error al crear técnico - no se pudo insertar en la base de datos', 'error')
            except Exception as e:
                print(f"Error creating technician: {e}")
                flash(f'Error al crear técnico: {str(e)}', 'error')
            
            return redirect(url_for('gestionar_tecnicos'))
        
        elif 'edit_tecnico' in request.form:
            tecnico_id = request.form.get('tecnico_id')
            if not tecnico_id:
                flash('ID de técnico inválido', 'error')
                return redirect(url_for('gestionar_tecnicos'))
            
            update_data = {
                'nombre': request.form.get('edit_nombre', '').strip(),
                'apellido': request.form.get('edit_apellido', '').strip(),
                'username': request.form.get('edit_username', '').strip(),
                'email': request.form.get('edit_email', '').strip()
            }
            
            if not all(update_data.values()):
                flash('Todos los campos son requeridos', 'error')
                return redirect(url_for('gestionar_tecnicos'))
            
            if request.form.get('edit_password'):
                update_data['password'] = bcrypt.hashpw(request.form['edit_password'].encode('utf-8'), bcrypt.gensalt())
            
            try:
                result = users.update_one({'_id': ObjectId(tecnico_id)}, {'$set': update_data})
                if result.modified_count > 0:
                    log_action('UPDATE_TECNICO', f'Técnico {update_data["username"]} actualizado')
                    flash('Técnico actualizado exitosamente', 'success')
                else:
                    flash('No se realizaron cambios', 'warning')
            except Exception as e:
                flash(f'Error al actualizar técnico: {str(e)}', 'error')
            
            return redirect(url_for('gestionar_tecnicos'))
        
        elif 'delete_tecnico' in request.form:
            tecnico_id = request.form.get('tecnico_id')
            if not tecnico_id:
                flash('ID de técnico inválido', 'error')
                return redirect(url_for('gestionar_tecnicos'))
            
            try:
                # Verificar si el técnico tiene bienes asignados
                if bienes_asignados.find_one({'tecnico_id': ObjectId(tecnico_id), 'estado': 'activo'}):
                    flash('No se puede eliminar: el técnico tiene bienes asignados', 'error')
                    return redirect(url_for('gestionar_tecnicos'))
                
                tecnico = users.find_one({'_id': ObjectId(tecnico_id)})
                result = users.delete_one({'_id': ObjectId(tecnico_id)})
                if result.deleted_count > 0:
                    log_action('DELETE_TECNICO', f'Técnico {tecnico.get("username", "")} eliminado')
                    flash('Técnico eliminado exitosamente', 'success')
                else:
                    flash('Error al eliminar técnico', 'error')
            except Exception as e:
                flash(f'Error al eliminar técnico: {str(e)}', 'error')
            
            return redirect(url_for('gestionar_tecnicos'))
    
    # Filtrar por parroquia si no es super admin
    if current_user.role == 'super_admin':
        tecnicos_list = list(users.find({'role': 'tecnico'}))
        parroquia_info = {'nombre': 'Sistema CONA', 'canton': 'Ecuador'}
        stats = {
            'total_bienes': inventarios.count_documents({}),
            'bienes_disponibles': inventarios.count_documents({'estado': 'disponible'}),
            'bienes_asignados': inventarios.count_documents({'estado': 'asignado'}),
            'bienes_mantenimiento': inventarios.count_documents({'estado': 'en_mantenimiento'}),
            'total_tecnicos': len(tecnicos_list)
        }
    else:
        filter_query = {'role': 'tecnico', 'parroquia_id': ObjectId(current_user.parroquia_id)} if current_user.parroquia_id else {'role': 'tecnico'}
        tecnicos_list = list(users.find(filter_query))
        parroquia_info = parroquias.find_one({'_id': ObjectId(current_user.parroquia_id)}) if current_user.parroquia_id else {'nombre': 'Sistema CONA', 'canton': 'Ecuador'}
        
        # Calcular estadísticas para la parroquia específica
        if current_user.parroquia_id:
            parroquia_filter = {'parroquia_id': ObjectId(current_user.parroquia_id)}
            stats = {
                'total_bienes': inventarios.count_documents(parroquia_filter),
                'bienes_disponibles': inventarios.count_documents({**parroquia_filter, 'estado': 'disponible'}),
                'bienes_asignados': inventarios.count_documents({**parroquia_filter, 'estado': 'asignado'}),
                'bienes_mantenimiento': inventarios.count_documents({**parroquia_filter, 'estado': 'en_mantenimiento'}),
                'total_tecnicos': len(tecnicos_list)
            }
        else:
            stats = {
                'total_bienes': inventarios.count_documents({}),
                'bienes_disponibles': inventarios.count_documents({'estado': 'disponible'}),
                'bienes_asignados': inventarios.count_documents({'estado': 'asignado'}),
                'bienes_mantenimiento': inventarios.count_documents({'estado': 'en_mantenimiento'}),
                'total_tecnicos': len(tecnicos_list)
            }
    
    return render_template('gestionar_tecnicos.html', tecnicos=tecnicos_list, parroquia=parroquia_info, stats=stats)

@app.route('/gestionar_asignaciones', methods=['GET', 'POST'])
@login_required
def gestionar_asignaciones():
    if request.method == 'POST':
        if 'assign_bien' in request.form:
            bien_id = request.form.get('bien_id')
            tecnico_id = request.form.get('tecnico_id')
            observaciones = request.form.get('observaciones', '')
            
            if not bien_id or not tecnico_id:
                flash('Debe seleccionar un bien y un técnico', 'error')
                return redirect(url_for('gestionar_asignaciones'))
            
            try:
                # Verificar que el bien esté disponible
                bien = inventarios.find_one({'_id': ObjectId(bien_id), 'estado': 'disponible'})
                if not bien:
                    flash('El bien no está disponible para asignación', 'error')
                    return redirect(url_for('gestionar_asignaciones'))
                
                asignacion = {
                    'bien_id': ObjectId(bien_id),
                    'tecnico_id': ObjectId(tecnico_id),
                    'parroquia_id': ObjectId(current_user.parroquia_id) if current_user.parroquia_id else None,
                    'fecha_asignacion': datetime.now(),
                    'observaciones': observaciones,
                    'estado': 'activo',
                    'created_at': datetime.now()
                }
                
                bienes_asignados.insert_one(asignacion)
                inventarios.update_one({'_id': ObjectId(bien_id)}, {'$set': {'estado': 'asignado'}})
                log_action('ASSIGN_BIEN', f'Bien {bien.get("codigo")} asignado')
                flash('Bien asignado exitosamente', 'success')
            except Exception as e:
                flash(f'Error al asignar bien: {str(e)}', 'error')
        elif 'devolver_bien' in request.form:
            asignacion_id = request.form.get('asignacion_id')
            bien_id = request.form.get('bien_id')
            try:
                bienes_asignados.update_one({'_id': ObjectId(asignacion_id)}, {'$set': {'estado': 'devuelto', 'fecha_devolucion': datetime.now()}})
                inventarios.update_one({'_id': ObjectId(bien_id)}, {'$set': {'estado': 'disponible'}})
                log_action('RETURN_BIEN', f'Bien devuelto')
                flash('Bien devuelto exitosamente', 'success')
            except Exception as e:
                flash(f'Error al devolver bien: {str(e)}', 'error')
    
    # Filtrar por parroquia si no es super admin
    if current_user.role == 'super_admin':
        bienes_disponibles = list(inventarios.find({'estado': 'disponible'}))
        tecnicos_list = list(users.find({'role': 'tecnico'}))
        parroquia_info = {'nombre': 'Sistema CONA', 'canton': 'Ecuador'}
        asignacion_filter = {'estado': 'activo'}
    else:
        parroquia_filter = ObjectId(current_user.parroquia_id) if current_user.parroquia_id else None
        if parroquia_filter:
            bienes_disponibles = list(inventarios.find({'estado': 'disponible', 'parroquia_id': parroquia_filter}))
            tecnicos_list = list(users.find({'role': 'tecnico', 'parroquia_id': parroquia_filter}))
            asignacion_filter = {'estado': 'activo', 'parroquia_id': parroquia_filter}
        else:
            bienes_disponibles = list(inventarios.find({'estado': 'disponible'}))
            tecnicos_list = list(users.find({'role': 'tecnico'}))
            asignacion_filter = {'estado': 'activo'}
        parroquia_info = parroquias.find_one({'_id': parroquia_filter}) if parroquia_filter else {'nombre': 'Sistema CONA', 'canton': 'Ecuador'}
    
    asignaciones_list = []
    for asignacion in bienes_asignados.find(asignacion_filter):
        bien_info = inventarios.find_one({'_id': asignacion.get('bien_id')})
        tecnico_info = users.find_one({'_id': asignacion.get('tecnico_id')}) if asignacion.get('tecnico_id') else None
        asignacion['bien_info'] = bien_info or {'codigo': 'N/A', 'nombre': 'N/A'}
        asignacion['tecnico_info'] = tecnico_info or {'nombre': 'N/A', 'apellido': ''}
        asignaciones_list.append(asignacion)
    
    stats = {
        'total_bienes': len(bienes_disponibles),
        'bienes_disponibles': inventarios.count_documents({'estado': 'disponible'}),
        'bienes_asignados': inventarios.count_documents({'estado': 'asignado'}),
        'bienes_mantenimiento': inventarios.count_documents({'estado': 'en_mantenimiento'}),
        'total_tecnicos': len(tecnicos_list)
    }
    
    return render_template('gestionar_asignaciones.html', bienes=bienes_disponibles, tecnicos=tecnicos_list, asignaciones=asignaciones_list, parroquia=parroquia_info, stats=stats)

@app.route('/asignar_bienes')
@login_required
def asignar_bienes():
    return render_template('asignar_bienes.html')

@app.route('/estadisticas')
@login_required
def estadisticas():
    if current_user.role != 'super_admin':
        return redirect(url_for('index'))
    
    # Calcular bienes por tipo
    pipeline = [
        {'$group': {'_id': '$tipo', 'count': {'$sum': 1}}},
        {'$sort': {'count': -1}}
    ]
    tipos_result = list(inventarios.aggregate(pipeline))
    bienes_por_tipo = []
    for tipo in tipos_result:
        if tipo['_id']:
            bienes_por_tipo.append({
                '_id': tipo['_id'],
                'cantidad': tipo['count']
            })
    
    stats = {
        'total_parroquias': parroquias.count_documents({}),
        'total_usuarios': users.count_documents({}),
        'total_bienes': inventarios.count_documents({}),
        'bienes_asignados': bienes_asignados.count_documents({'estado': 'activo'}),
        'total_tecnicos': users.count_documents({'role': 'tecnico'}),
        'asignaciones_activas': bienes_asignados.count_documents({'estado': 'activo'}),
        'asignaciones_devueltas': bienes_asignados.count_documents({'estado': 'devuelto'}),
        'parroquias_con_bienes': parroquias.count_documents({}),
        'responsables_activos': users.count_documents({'role': {'$in': ['admin_parroquia', 'tecnico']}}),
        'bienes_por_estado': {
            'disponibles': inventarios.count_documents({'estado': 'disponible'}),
            'asignados': inventarios.count_documents({'estado': 'asignado'}),
            'mantenimiento': inventarios.count_documents({'estado': 'en_mantenimiento'}),
            'dañados': inventarios.count_documents({'estado': 'dañado'})
        },
        'usuarios_por_rol': {
            'super_admin': users.count_documents({'role': 'super_admin'}),
            'admin_parroquia': users.count_documents({'role': 'admin_parroquia'}),
            'tecnico': users.count_documents({'role': 'tecnico'})
        },
        'bienes_por_tipo': bienes_por_tipo
    }
    
    return render_template('estadisticas.html', stats=stats, estadisticas=stats)

@app.route('/estadisticas_parroquia')
@login_required
def estadisticas_parroquia():
    # Calcular bienes por tipo para la parroquia
    if current_user.parroquia_id:
        pipeline = [
            {'$match': {'parroquia_id': ObjectId(current_user.parroquia_id)}},
            {'$group': {'_id': '$tipo', 'count': {'$sum': 1}}},
            {'$sort': {'count': -1}}
        ]
    else:
        pipeline = [
            {'$group': {'_id': '$tipo', 'count': {'$sum': 1}}},
            {'$sort': {'count': -1}}
        ]
    
    tipos_result = list(inventarios.aggregate(pipeline))
    bienes_por_tipo = []
    for tipo in tipos_result:
        if tipo['_id']:
            bienes_por_tipo.append({
                '_id': tipo['_id'],
                'cantidad': tipo['count']
            })
    
    stats = {
        'total_bienes': inventarios.count_documents({'parroquia_id': ObjectId(current_user.parroquia_id)}) if current_user.parroquia_id else 0,
        'bienes_disponibles': inventarios.count_documents({'parroquia_id': ObjectId(current_user.parroquia_id), 'estado': 'disponible'}) if current_user.parroquia_id else 0,
        'bienes_asignados': inventarios.count_documents({'parroquia_id': ObjectId(current_user.parroquia_id), 'estado': 'asignado'}) if current_user.parroquia_id else 0,
        'bienes_mantenimiento': inventarios.count_documents({'parroquia_id': ObjectId(current_user.parroquia_id), 'estado': 'en_mantenimiento'}) if current_user.parroquia_id else 0,
        'total_tecnicos': users.count_documents({'role': 'tecnico', 'parroquia_id': ObjectId(current_user.parroquia_id)}) if current_user.parroquia_id else 0,
        'bienes_por_tipo': bienes_por_tipo
    }
    
    parroquia_info = {'nombre': 'Sistema CONA', 'canton': 'Ecuador'}
    return render_template('estadisticas_parroquia.html', stats=stats, estadisticas=stats, parroquia=parroquia_info)

@app.route('/auditoria')
@login_required
def auditoria():
    if current_user.role != 'super_admin':
        return redirect(url_for('index'))
    
    logs = list(audit_logs.find().sort('timestamp', -1).limit(100))
    return render_template('auditoria.html', logs=logs)

# Rutas para reportes
@app.route('/reporte_pdf/<tipo>')
@login_required
def generar_reporte_pdf(tipo):
    try:
        from reportlab.lib.pagesizes import A4, letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        buffer = BytesIO()
        # Usar formato A4 horizontal para más espacio
        from reportlab.lib.pagesizes import landscape
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=50, bottomMargin=30)
        
        styles = getSampleStyleSheet()
        story = []
        
        # Header
        header_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=20, alignment=1)
        title = Paragraph(f"REPORTE DE {tipo.upper()}", header_style)
        story.append(title)
        
        # Info
        info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, spaceAfter=15)
        fecha = datetime.now().strftime('%d/%m/%Y %H:%M')
        info = Paragraph(f"Generado el: {fecha} | Usuario: {current_user.username}", info_style)
        story.append(info)
        
        story.append(Spacer(1, 15))
        
        # Función para ajustar texto sin cortar
        def ajustar_texto(texto, max_chars=None):
            if not texto:
                return ''
            return str(texto)
        
        # Datos según tipo
        if tipo == 'inventario':
            if current_user.role == 'super_admin':
                bienes = list(inventarios.find())
            else:
                filter_query = {'parroquia_id': ObjectId(current_user.parroquia_id)} if current_user.parroquia_id else {}
                bienes = list(inventarios.find(filter_query))
            
            from reportlab.platypus import Image as RLImage
            
            # Crear estilo blanco para header
            white_style = ParagraphStyle('WhiteHeader', parent=styles['Normal'], textColor=colors.whitesmoke, fontSize=10, alignment=1)
            
            # Headers con foto incluida (sin fecha)
            data = [[
                'Foto',
                'Código', 
                'Nombre',
                'Tipo',
                'Marca/Modelo',
                'Color',
                'Estado Equipo',
                Paragraph('Estado<br/>Asignación', white_style)
            ]]
            
            for bien in bienes:
                # Procesar imagen
                imagen_cell = ''
                if bien.get('imagen'):
                    try:
                        img_data = base64.b64decode(bien['imagen'])
                        img_buffer = BytesIO(img_data)
                        rl_img = RLImage(img_buffer, width=0.8*inch, height=0.8*inch)
                        imagen_cell = rl_img
                    except:
                        imagen_cell = Paragraph('Sin imagen', styles['Normal'])
                else:
                    imagen_cell = Paragraph('Sin imagen', styles['Normal'])
                
                # Usar Paragraph para texto que se ajuste con saltos de línea
                marca_modelo = f"{bien.get('marca', '')} {bien.get('modelo', '')}".strip()
                
                data.append([
                    imagen_cell,
                    Paragraph(str(bien.get('codigo', '')), styles['Normal']),
                    Paragraph(str(bien.get('nombre', '')), styles['Normal']),
                    Paragraph(str(bien.get('tipo', '')), styles['Normal']),
                    Paragraph(str(marca_modelo), styles['Normal']),
                    Paragraph(str(bien.get('color', 'N/A')), styles['Normal']),
                    Paragraph(str(bien.get('estado_equipo', 'N/A')), styles['Normal']),
                    Paragraph(str(bien.get('estado', 'disponible')), styles['Normal'])
                ])
            
            # Anchos optimizados con foto incluida (sin fecha)
            col_widths = [1*inch, 0.8*inch, 1.8*inch, 0.8*inch, 2*inch, 0.8*inch, 1.3*inch, 1.3*inch]
        
        elif tipo == 'asignaciones':
            if current_user.role == 'super_admin':
                asignaciones = list(bienes_asignados.find({'estado': 'activo'}))
            else:
                filter_query = {'estado': 'activo', 'parroquia_id': ObjectId(current_user.parroquia_id)} if current_user.parroquia_id else {'estado': 'activo'}
                asignaciones = list(bienes_asignados.find(filter_query))
            
            data = [['Bien', 'Técnico', 'Fecha', 'Observaciones']]
            for asignacion in asignaciones:
                bien_info = inventarios.find_one({'_id': asignacion.get('bien_id')})
                tecnico_info = users.find_one({'_id': asignacion.get('tecnico_id')})
                
                bien_texto = f"{bien_info.get('codigo', 'N/A')} - {bien_info.get('nombre', 'N/A')}" if bien_info else 'N/A'
                tecnico_texto = f"{tecnico_info.get('nombre', 'N/A')} {tecnico_info.get('apellido', '')}" if tecnico_info else 'N/A'
                fecha_texto = asignacion.get('fecha_asignacion', '').strftime('%d/%m/%Y') if asignacion.get('fecha_asignacion') else 'N/A'
                
                data.append([
                    ajustar_texto(bien_texto, 22),
                    ajustar_texto(tecnico_texto, 18),
                    fecha_texto,
                    ajustar_texto(asignacion.get('observaciones', ''), 25)
                ])
            
            # Anchos de columna ajustados
            col_widths = [2.2*inch, 1.8*inch, 1*inch, 2.5*inch]
        
        # Crear tabla con altura variable para ajuste de texto
        table = Table(data, colWidths=col_widths, repeatRows=1, rowHeights=None)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('WORDWRAP', (0, 0), (-1, -1), 'LTR')
        ]))
        
        story.append(table)
        doc.build(story)
        
        pdf = buffer.getvalue()
        buffer.close()
        
        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="reporte_{tipo}_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
        
        return response
        
    except ImportError as ie:
        flash('Librería ReportLab no instalada', 'error')
        return redirect(request.referrer or url_for('index'))
    except Exception as e:
        print(f'Error generando PDF: {str(e)}')
        flash(f'Error generando PDF: {str(e)}', 'error')
        return redirect(request.referrer or url_for('index'))

@app.route('/exportar_excel/<tipo>')
@login_required
def exportar_excel(tipo):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Reporte {tipo.title()}"
        
        if tipo == 'inventario':
            if current_user.role == 'super_admin':
                bienes = list(inventarios.find())
            else:
                filter_query = {'parroquia_id': ObjectId(current_user.parroquia_id)} if current_user.parroquia_id else {}
                bienes = list(inventarios.find(filter_query))
            
            headers = ['Código', 'Nombre', 'Tipo', 'Marca', 'Modelo', 'Estado', 'Fecha Creación']
            ws.append(headers)
            
            # Estilo para headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for bien in bienes:
                ws.append([
                    bien.get('codigo', ''),
                    bien.get('nombre', ''),
                    bien.get('tipo', ''),
                    bien.get('marca', ''),
                    bien.get('modelo', ''),
                    bien.get('estado', ''),
                    bien.get('created_at', '').strftime('%d/%m/%Y') if bien.get('created_at') else ''
                ])
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="reporte_{tipo}.xlsx"'
        
        return response
        
    except ImportError:
        flash('Librería openpyxl no instalada', 'error')
        return redirect(request.referrer or url_for('index'))
    except Exception as e:
        flash(f'Error generando Excel: {str(e)}', 'error')
        return redirect(request.referrer or url_for('index'))

@app.route('/generar_qr/<bien_id>')
@login_required
def generar_qr(bien_id):
    if not QR_AVAILABLE:
        flash('Función de generar QR no disponible', 'warning')
        return redirect(url_for('index'))
    
    try:
        bien = inventarios.find_one({'_id': ObjectId(bien_id)})
        if not bien:
            flash('Bien no encontrado', 'error')
            return redirect(url_for('gestionar_inventario'))
        
        # Generar QR con información del bien
        qr_data = f"Código: {bien.get('codigo')}\nNombre: {bien.get('nombre')}\nTipo: {bien.get('tipo')}\nMarca: {bien.get('marca')}\nModelo: {bien.get('modelo')}"
        
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Convertir a base64 para mostrar en el navegador
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        img_str = base64.b64encode(buffer.getvalue()).decode()
        
        return render_template('qr_code.html', qr_image=img_str, bien=bien)
    except Exception as e:
        flash(f'Error al generar QR: {str(e)}', 'error')
        return redirect(url_for('gestionar_inventario'))

@app.route('/panel_tecnico')
@login_required
def panel_tecnico():
    if current_user.role != 'tecnico':
        return redirect(url_for('index'))
    
    # Obtener información del técnico
    user_data = users.find_one({'username': current_user.username})
    if not user_data:
        return redirect(url_for('login'))
    
    # Buscar parroquia del técnico
    parroquia_info = None
    if current_user.parroquia_id:
        parroquia_info = parroquias.find_one({'_id': ObjectId(current_user.parroquia_id)})
    
    if not parroquia_info:
        parroquia_info = {'nombre': 'Parroquia Demo', 'canton': 'Canton Demo'}
    
    # Obtener bienes asignados al técnico
    bienes_asignados_tecnico = []
    asignaciones = list(bienes_asignados.find({
        'tecnico_id': ObjectId(user_data['_id']),
        'estado': 'activo'
    }))
    
    for asignacion in asignaciones:
        bien_info = inventarios.find_one({'_id': asignacion.get('bien_id')})
        if bien_info:
            bien_info['fecha_asignacion'] = asignacion.get('fecha_asignacion')
            bien_info['observaciones'] = asignacion.get('observaciones', '')
            bienes_asignados_tecnico.append(bien_info)
    
    # Estadísticas del técnico
    stats = {
        'total_bienes_asignados': len(bienes_asignados_tecnico),
        'bienes_por_tipo': {},
        'fecha_ultima_asignacion': None
    }
    
    # Contar bienes por tipo
    for bien in bienes_asignados_tecnico:
        tipo = bien.get('tipo', 'Sin especificar')
        stats['bienes_por_tipo'][tipo] = stats['bienes_por_tipo'].get(tipo, 0) + 1
    
    # Fecha de última asignación
    if asignaciones:
        fechas = [a.get('fecha_asignacion') for a in asignaciones if a.get('fecha_asignacion')]
        if fechas:
            stats['fecha_ultima_asignacion'] = max(fechas)
    
    return render_template('panel_tecnico.html', 
                         parroquia=parroquia_info, 
                         tecnico=user_data,
                         bienes=bienes_asignados_tecnico,
                         stats=stats)

@app.route('/mi_cuenta', methods=['GET', 'POST'])
@login_required
def mi_cuenta():
    if request.method == 'POST':
        if 'cambiar_password' in request.form:
            password_actual = request.form['password_actual']
            password_nuevo = request.form['password_nuevo']
            confirmar_password = request.form['confirmar_password']
            
            user = users.find_one({'username': current_user.username})
            if user and bcrypt.checkpw(password_actual.encode('utf-8'), user['password'].encode('utf-8') if isinstance(user['password'], str) else user['password']):
                if password_nuevo == confirmar_password:
                    hashed_password = bcrypt.hashpw(password_nuevo.encode('utf-8'), bcrypt.gensalt())
                    users.update_one({'username': current_user.username}, {'$set': {'password': hashed_password}})
                    flash('Contraseña actualizada exitosamente', 'success')
                else:
                    flash('Las contraseñas no coinciden', 'error')
            else:
                flash('Contraseña actual incorrecta', 'error')
    
    user_data = users.find_one({'username': current_user.username})
    if not user_data:
        user_data = {'username': current_user.username, 'nombre': '', 'apellido': '', 'email': ''}
    return render_template('mi_cuenta.html', user=user_data, user_data=user_data)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 10000))
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(host='0.0.0.0', port=port, debug=debug)